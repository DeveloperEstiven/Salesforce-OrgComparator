from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

missing_fill_org1 = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
missing_fill_org2 = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
header_font = Font(bold=True)

# Caching describe results for each object to avoid repeated API calls
describe_cache = {}

def get_describe_result(org, obj_name):
    if (org, obj_name) not in describe_cache:
        describe_cache[(org, obj_name)] = org.__getattr__(obj_name).describe()
    return describe_cache[(org, obj_name)]

def get_field_info(org, obj_name, field_name):
    describe_result = get_describe_result(org, obj_name)
    fields_info = {field['name']: field for field in describe_result['fields']}
    field_info = fields_info.get(field_name, {})
    field_type = field_info.get('type', 'Unknown')
    field_label = field_info.get('label', field_name)
    return field_type, field_label

def write_sorted_fields(ws, obj_name, sorted_fields, org1, org2, org1_name, org2_name):
    for field_name, org1_exists, org2_exists in sorted_fields:
        field_type_org1, field_label_org1 = get_field_info(org1, obj_name, field_name) if org1_exists else ('', '')
        field_type_org2, field_label_org2 = get_field_info(org2, obj_name, field_name) if org2_exists else ('', '')
        field_type = field_type_org1 or field_type_org2
        field_label = field_label_org1 or field_label_org2
        
        # Update status message with org names
        status = f"Field missing in {org2_name}" if org1_exists and not org2_exists else f"Field missing in {org1_name}" if not org1_exists and org2_exists else "Field exists in both orgs"
        
        # Updated row structure: Field Name in second column, Field Label in third, Field Type in fourth
        row = [obj_name, field_name, field_label, field_type, 'Yes' if org1_exists else 'No', 'Yes' if org2_exists else 'No', status]
        ws.append(row)
        
        if not org1_exists:
            for cell in ws[ws.max_row]:
                cell.fill = missing_fill_org1
        elif not org2_exists:
            for cell in ws[ws.max_row]:
                cell.fill = missing_fill_org2

def compare_and_write_sheet(ws, org1_objects, org2_objects, org1, org2, org1_name, org2_name):
    org1_object_names = {obj['name']: obj for obj in org1_objects}
    org2_object_names = {obj['name']: obj for obj in org2_objects}
    
    # Update column headers with the correct order
    ws.append(['Object', 'Field Name', 'Field Label', 'Field Type', f'{org1_name} Field Exists', f'{org2_name} Field Exists', 'Status'])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    # Process each object and display its name
    for obj_name, org1_obj in org1_object_names.items():
        print(f"Processing object: {obj_name}")
        org1_fields = {field['name'] for field in get_describe_result(org1, obj_name)['fields']}
        org2_fields = {field['name'] for field in get_describe_result(org2, obj_name)['fields']} if obj_name in org2_object_names else set()
        sorted_fields = sorted([(field, field in org1_fields, field in org2_fields) for field in org1_fields.union(org2_fields)], key=lambda x: (not x[1], not x[2]))
        write_sorted_fields(ws, obj_name, sorted_fields, org1, org2, org1_name, org2_name)

    for obj_name in org2_object_names:
        if obj_name not in org1_object_names:
            print(f"Processing object: {obj_name} (missing in {org1_name})")
            org2_fields = {field['name'] for field in get_describe_result(org2, obj_name)['fields']}
            sorted_fields = [(field, False, True) for field in org2_fields]
            write_sorted_fields(ws, obj_name, sorted_fields, org1, org2, org1_name, org2_name)

    print("Comparison completed.")
