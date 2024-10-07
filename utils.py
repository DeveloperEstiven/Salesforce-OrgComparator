from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_dropdowns(ws, max_row, cols, org1_name, org2_name):
    yes_no_dv = DataValidation(type="list", formula1='"Yes,No"', showDropDown=True)

    # Dynamically create the status dropdown based on org1 and org2 names
    status_dv = DataValidation(
        type="list",
        formula1=f'"Field exists in both orgs,Field missing in {org1_name},Field missing in {org2_name}"',
        showDropDown=True
    )

    ws.add_data_validation(yes_no_dv)
    ws.add_data_validation(status_dv)

    # Apply dropdowns to the appropriate columns
    for row in range(2, max_row + 1):
        yes_no_dv.add(ws[f'{cols[0]}{row}'])  # Dropdown for Org1 Field Exists
        yes_no_dv.add(ws[f'{cols[1]}{row}'])  # Dropdown for Org2 Field Exists
        status_dv.add(ws[f'{cols[2]}{row}'])  # Dropdown for Status

def auto_adjust_column_widths(ws):
    """Automatically adjust the width of each column based on its content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)  # Add padding for readability
        ws.column_dimensions[col_letter].width = adjusted_width
